import os
import time
import asyncio
import logging
from io import BytesIO
from typing import List

logger = logging.getLogger(__name__)

CHUNK_SIZE = 1500
CHUNK_OVERLAP = 150
SUPPORTED_EXTS = {
    ".md", ".txt", ".pdf", ".docx", ".doc",
    ".csv", ".json", ".yaml", ".toml",
    ".py", ".js", ".ts", ".html", ".css", ".sh",
    ".xlsx", ".xls", ".pptx", ".ppt",
    ".eml", ".msg", ".rtf", ".odt",
}
MIN_CONTENT_LEN = 50

COLLECTION_KNOWLEDGE = "jkai_knowledge"
COLLECTION_EXTERNAL = "jkai_external"


def _parse_bytes(file_bytes: bytes, filename: str) -> str:
    """
    Parse raw bytes of a file into plain text.
    Uses unstructured for rich documents (PDF, DOCX, XLSX, PPTX, ...),
    falls back to plain text decoding for text-based formats.
    Does NOT write anything to disk — operates entirely on BytesIO in RAM.
    """
    ext = os.path.splitext(filename)[1].lower()
    plaintext_exts = {
        ".md", ".txt", ".py", ".js", ".ts", ".html", ".css", ".sh",
        ".json", ".yaml", ".toml", ".csv",
    }

    if ext in plaintext_exts:
        try:
            return file_bytes.decode("utf-8", errors="replace")
        except Exception:
            return ""

    # Use unstructured for rich document types — reads from BytesIO, no temp file needed
    try:
        from unstructured.partition.auto import partition

        bio = BytesIO(file_bytes)
        # metadata_filename tells unstructured which parser to use based on extension
        elements = partition(file=bio, metadata_filename=filename)
        parts = []
        for el in elements:
            text = str(el).strip()
            if text:
                parts.append(text)
        return "\n\n".join(parts)

    except ImportError:
        logger.warning("unstructured not available, falling back to basic parsers")
        return _parse_bytes_fallback(file_bytes, filename)
    except Exception as e:
        logger.warning(f"unstructured failed for {filename}: {e}, falling back")
        return _parse_bytes_fallback(file_bytes, filename)


def _parse_bytes_fallback(file_bytes: bytes, filename: str) -> str:
    """Basic fallback parsers using BytesIO — still no temp file written."""
    ext = os.path.splitext(filename)[1].lower()
    bio = BytesIO(file_bytes)
    try:
        if ext == ".docx":
            from docx import Document
            doc = Document(bio)
            parts = []
            for el in doc.element.body:
                tag = el.tag.split("}")[-1] if "}" in el.tag else el.tag
                if tag == "p":
                    from docx.oxml.ns import qn
                    text = "".join(n.text or "" for n in el.iter(qn("w:t")))
                    if text.strip():
                        parts.append(text.strip())
                elif tag == "tbl":
                    # Extract table rows
                    for row in el.iter(el.tag.split("}")[0] + "}tr"):
                        cells = []
                        for cell in row.iter(el.tag.split("}")[0] + "}tc"):
                            cell_text = "".join(n.text or "" for n in cell.iter())
                            cells.append(cell_text.strip())
                        if any(cells):
                            parts.append(" | ".join(cells))
            return "\n".join(parts)

        elif ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        lines.append(row_text)
            return "\n".join(lines)

        elif ext == ".pdf":
            import pdfplumber
            with pdfplumber.open(bio) as pdf:
                return "\n".join(page.extract_text() or "" for page in pdf.pages)

    except Exception as e:
        logger.warning(f"Fallback parse failed for {filename}: {e}")

    return ""


def _chunk_text(text: str) -> List[dict]:
    """
    Split text into overlapping chunks, preferring paragraph boundaries.
    Tries to split at double-newlines (paragraph breaks) first,
    then single newlines, then raw character position.
    Returns a list of dicts with text and character offsets.
    """
    text = text.strip()
    if not text:
        return []
    if len(text) <= CHUNK_SIZE:
        return [{"text": text, "start_char_idx": 0, "end_char_idx": len(text)}]

    chunks = []
    start = 0
    while start < len(text):
        end = min(start + CHUNK_SIZE, len(text))
        if end < len(text):
            # Prefer paragraph break
            para_pos = text.rfind("\n\n", start, end)
            if para_pos > start + CHUNK_OVERLAP:
                end = para_pos
            else:
                # Fallback: single newline
                nl_pos = text.rfind("\n", start, end)
                if nl_pos > start + CHUNK_OVERLAP:
                    end = nl_pos
        chunk = text[start:end].strip()
        if chunk:
            chunks.append({
                "text": chunk,
                "start_char_idx": start,
                "end_char_idx": end
            })
        start = end - CHUNK_OVERLAP if end < len(text) else end
    return chunks


class IngestionPipeline:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self):
        if self._initialized:
            return
        self._initialized = True

    async def ingest_directory(
        self,
        directory: str,
        source_id: str = "local",
        file_type_tag: str = "docs",
        collection: str = None,
        task_id: str = "ingest",
    ) -> dict:
        from core.qdrant_client import qdrant_client
        from core.utils.embed import embed
        from core.knowledge_sources.metadata import metadata_db

        target_collection = collection or COLLECTION_KNOWLEDGE
        stats = {"scanned": 0, "new": 0, "updated": 0, "failed": 0, "skipped": 0}

        if not os.path.isdir(directory):
            return {**stats, "error": f"Directory not found: {directory}"}

        metadata_db.register_source(
            source_id, os.path.basename(directory), "local", {"path": directory}
        )

        all_points_buffer = []

        for root, _, files in os.walk(directory):
            for filename in sorted(files):
                ext = os.path.splitext(filename)[1].lower()
                if ext not in SUPPORTED_EXTS:
                    continue
                filepath = os.path.join(root, filename)
                rel_path = os.path.relpath(filepath, directory)
                stats["scanned"] += 1

                # --- Change detection: mtime + size only (NO full-file read) ---
                try:
                    stat = os.stat(filepath)
                    mtime = stat.st_mtime
                    fsize = stat.st_size
                except OSError as e:
                    logger.warning(f"Cannot stat {filepath}: {e}")
                    stats["failed"] += 1
                    continue

                existing = metadata_db.get_file(source_id, rel_path)
                if (
                    existing
                    and existing.get("status") == "indexed"
                    and abs(existing.get("mtime", 0) - mtime) < 1.0
                    and existing.get("file_size", -1) == fsize
                ):
                    stats["skipped"] += 1
                    continue

                # --- Read file bytes once (Rclone VFS streams on-demand) ---
                try:
                    with open(filepath, "rb") as f:
                        file_bytes = f.read()
                except Exception as e:
                    logger.warning(f"Cannot read {filepath}: {e}")
                    metadata_db.mark_failed(source_id, rel_path, str(e))
                    stats["failed"] += 1
                    continue

                # --- Parse in-memory from bytes (no temp file) ---
                content = _parse_bytes(file_bytes, filename)
                del file_bytes  # release RAM immediately after parsing

                if not content or len(content.strip()) < MIN_CONTENT_LEN:
                    stats["skipped"] += 1
                    continue

                chunks = _chunk_text(content)
                if not chunks:
                    stats["skipped"] += 1
                    continue

                # --- Embed and upsert to Qdrant ---
                try:
                    # Song song hóa việc gọi embedding cho các chunks trong cùng một file qua Semaphore thưa Master
                    sem = asyncio.Semaphore(3)

                    async def _embed_one(chunk_text):
                        async with sem:
                            try:
                                return await embed.get_embedding_async(chunk_text[:4000])
                            except Exception:
                                return None

                    vectors = await asyncio.gather(*[_embed_one(c["text"]) for c in chunks])

                    import uuid
                    file_points = []
                    for i, (chunk, vector) in enumerate(zip(chunks, vectors)):
                        if vector:
                            file_points.append({
                                "id": str(uuid.uuid5(uuid.NAMESPACE_DNS, f"{source_id}:{rel_path}:chunk_{i}")),
                                "vector": vector,
                                "payload": {
                                    "text": chunk["text"],
                                    "source": file_type_tag,
                                    "rel_path": rel_path,
                                    "source_path": filepath,
                                    "start_char_idx": chunk["start_char_idx"],
                                    "end_char_idx": chunk["end_char_idx"],
                                    "filename": filename,
                                    "file_type": ext,
                                    "chunk_index": i,
                                    "mtime": mtime,
                                    "file_size": fsize,
                                    "indexed_at": time.time(),
                                },
                            })

                    stats["failed"] += len(chunks) - len(file_points)
                    all_points_buffer.extend(file_points)

                    # Flush buffer khi gom đủ 200 points để tối ưu hóa Network round-trips
                    if len(all_points_buffer) >= 200:
                        await qdrant_client.upsert_batch(all_points_buffer[:200], target_collection)
                        all_points_buffer = all_points_buffer[200:]

                    # Store mtime+size in metadata
                    metadata_db.upsert_file(
                        source_id=source_id,
                        rel_path=rel_path,
                        abs_path=filepath,
                        file_type=ext,
                        checksum="",
                        file_size=fsize,
                        mtime=mtime,
                        status="indexed",
                    )

                    if existing:
                        stats["updated"] += 1
                    else:
                        stats["new"] += 1

                except Exception as e:
                    logger.error(f"Embed/upsert failed for {rel_path}: {e}")
                    metadata_db.mark_failed(source_id, rel_path, str(e))
                    stats["failed"] += 1

        # Flush lượng points còn lại trong buffer sau khi kết thúc vòng lặp thưa Master
        if all_points_buffer:
            try:
                await qdrant_client.upsert_batch(all_points_buffer, target_collection)
            except Exception as e:
                logger.error(f"Final batch flush failed: {e}")

        metadata_db._get_conn().execute(
            "UPDATE sources SET last_sync=?, updated_at=? WHERE id=?",
            (time.time(), time.time(), source_id),
        )
        metadata_db._get_conn().commit()

        return stats


ingestion_pipeline = IngestionPipeline()
