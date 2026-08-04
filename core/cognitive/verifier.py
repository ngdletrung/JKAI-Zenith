"""
JKAI ZENITH — COGNITIVE VERIFIER & RECOVERY ENGINE (v2.0)
File: core/cognitive/verifier.py

Thẩm định kết quả thực thi (Correct? Complete? Meets mission criteria?).
Nếu PASS -> DELIVER
Nếu FAIL -> RECOVER / REPLAN & đúc kết ExperienceRecord cho Engram v2.
"""

from __future__ import annotations
import os
import logging
from typing import Dict, Any, List, Optional

from core.contracts.cognitive_contract import (
    CognitiveRequest,
    DeliverableType,
    VerificationResult,
    ExperienceRecord,
)

logger = logging.getLogger("jkai.cognitive.verifier")


class CognitiveVerifier:
    """Bộ Thẩm Định Kết Quả Thực Thi (Cognitive Verifier)."""

    @classmethod
    def verify(cls, request: CognitiveRequest, result_payload: Dict[str, Any]) -> VerificationResult:
        """
        Thẩm định kết quả thực thi so với CognitiveRequest.
        """
        deliv = request.deliverable
        missing: List[str] = []
        logs: List[str] = []

        if deliv.type in (DeliverableType.FILE_BINARY, DeliverableType.FILE_CODE):
            target_path = result_payload.get("file_path") or deliv.target_path
            
            # 1. Kiểm tra tồn tại file trên đĩa
            if not target_path or not os.path.exists(target_path):
                missing.append(f"PHYSICAL_FILE_MISSING: {target_path}")
                logs.append(f"❌ File '{target_path}' does not exist on disk.")
            else:
                # 2. Kiểm tra dung lượng file > 0 bytes
                size = os.path.getsize(target_path)
                if size <= 0:
                    missing.append("FILE_EMPTY_ZERO_BYTES")
                    logs.append(f"❌ File '{target_path}' exists but is 0 bytes.")
                else:
                    logs.append(f"✅ Physical file '{target_path}' verified ({size} bytes).")

                # 3. Thẩm định định dạng file
                if deliv.format == "xlsx":
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(target_path, read_only=True)
                        logs.append(f"✅ Excel integrity verified: {len(wb.sheetnames)} sheets found ({wb.sheetnames}).")
                    except Exception as e:
                        missing.append(f"EXCEL_CORRUPTED: {e}")
                        logs.append(f"❌ Excel integrity check failed: {e}")

        passed = len(missing) == 0
        score = 1.0 if passed else max(0.0, 1.0 - (len(missing) * 0.4))

        ver_res = VerificationResult(
            passed=passed,
            score=score,
            summary="Verification PASSED 100%" if passed else f"Verification FAILED: {', '.join(missing)}",
            missing_criteria=missing,
            diagnostic_logs=logs
        )
        logger.info(f"🔎 [COGNITIVE-VERIFIER]: Passed={passed}, Score={score:.2f}, Summary='{ver_res.summary}'")
        return ver_res

    @classmethod
    def create_experience(
        cls,
        request: CognitiveRequest,
        verifier_result: VerificationResult,
        strategy: str,
        tools_used: List[str],
        failure_reason: Optional[str] = None
    ) -> ExperienceRecord:
        """
        Tạo Hồ sơ Trải nghiệm (ExperienceRecord) bao gồm cả Bài học tiêu cực (Negative Lessons) cho Engram v2.
        """
        negative_lessons: List[str] = []
        if not verifier_result.passed:
            for miss in verifier_result.missing_criteria:
                if "EXCEL_CORRUPTED" in miss:
                    negative_lessons.append("Avoid raw file stream writing for Excel. Use openpyxl.Workbook.save().")
                elif "PHYSICAL_FILE_MISSING" in miss:
                    negative_lessons.append("Ensure target directory exists before running script.")

        rec = ExperienceRecord(
            task_signature=f"{request.intent}_{request.deliverable.format}",
            context_summary=request.goal[:100],
            strategy_used=strategy,
            tools_used=tools_used,
            model_profile_used="AMG_SELECTED_MODEL",
            outcome="SUCCESS" if verifier_result.passed else "FAILED",
            failure_cause=failure_reason or verifier_result.summary if not verifier_result.passed else None,
            recovery_action="REPLAN_WITH_NEGATIVE_MEMORY" if not verifier_result.passed else None,
            verification_passed=verifier_result.passed,
            negative_lessons=negative_lessons,
            confidence_rating=verifier_result.score
        )
        logger.info(f"🧠 [ENGRAM-EXPERIENCE]: Logged record outcome={rec.outcome}, negative_lessons={len(negative_lessons)}")
        return rec
