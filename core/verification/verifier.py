"""
JKAI ZENITH — VERIFICATION PACKAGE: COGNITIVE VERIFIER (v2.1)
File: core/verification/verifier.py

Thẩm định kết quả thực thi (Correct? Complete? Meets mission criteria?).
Phân loại thất bại: FailureClassification (TRANSIENT, TOOL_FAILURE, MODEL_FAILURE, PLAN_FAILURE, VERIFICATION_FAILURE).
Đúc kết trải nghiệm Engram v2 với Bài học tiêu cực (Negative Memory qua từng Attempt).
"""

from __future__ import annotations
import os
import logging
from typing import Dict, Any, List, Optional

from core.contracts.cognitive_contract import CognitiveRequest, DeliverableType
from core.contracts.identity_contract import IdentityChain
from core.contracts.verification_contract import (
    VerificationResult,
    FailureClassification,
    RecoveryStrategy,
    ExperienceRecord,
)

logger = logging.getLogger("jkai.verification.verifier")


class CognitiveVerifier:
    """Bộ Thẩm Định Kết Quả Thực Thi (Cognitive Verifier v2.1)."""

    @classmethod
    def verify(
        cls,
        request: CognitiveRequest,
        result_payload: Dict[str, Any],
        identity: Optional[IdentityChain] = None
    ) -> VerificationResult:
        """
        Thẩm định kết quả thực thi so với CognitiveRequest.
        """
        deliv = request.deliverable
        ident = identity or request.identity
        missing: List[str] = []
        logs: List[str] = []
        fail_cls = FailureClassification.NONE
        rec_rec = RecoveryStrategy.NONE

        if deliv.type in (DeliverableType.FILE_BINARY, DeliverableType.FILE_CODE):
            target_path = result_payload.get("file_path") or deliv.target_path
            
            # 1. Kiểm tra tồn tại file trên đĩa
            if not target_path or not os.path.exists(target_path):
                missing.append(f"PHYSICAL_FILE_MISSING: {target_path}")
                logs.append(f"❌ File '{target_path}' does not exist on disk.")
                fail_cls = FailureClassification.TOOL_FAILURE
                rec_rec = RecoveryStrategy.SUBSTITUTE_CAPABILITY
            else:
                # 2. Kiểm tra dung lượng file > 0 bytes
                size = os.path.getsize(target_path)
                if size <= 0:
                    missing.append("FILE_EMPTY_ZERO_BYTES")
                    logs.append(f"❌ File '{target_path}' exists but is 0 bytes.")
                    fail_cls = FailureClassification.VERIFICATION_FAILURE
                    rec_rec = RecoveryStrategy.DIAGNOSE_AND_REPAIR
                else:
                    logs.append(f"✅ Physical file '{target_path}' verified ({size} bytes).")

                # 3. Thẩm định định dạng file
                if deliv.format == "xlsx":
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(target_path, read_only=True)
                        logs.append(f"✅ Excel integrity verified: {len(wb.sheetnames)} sheets found ({wb.sheetnames}).")
                    except Exception as e:
                        missing.append(f"EXCEL_CORRUPTED: {e}")
                        logs.append(f"❌ Excel integrity check failed: {e}")
                        fail_cls = FailureClassification.VERIFICATION_FAILURE
                        rec_rec = RecoveryStrategy.DIAGNOSE_AND_REPAIR

        passed = len(missing) == 0
        score = 1.0 if passed else max(0.0, 1.0 - (len(missing) * 0.4))

        ver_res = VerificationResult(
            identity=ident,
            passed=passed,
            score=score,
            failure_classification=fail_cls if not passed else FailureClassification.NONE,
            recommended_recovery=rec_rec if not passed else RecoveryStrategy.NONE,
            summary="Verification PASSED 100%" if passed else f"Verification FAILED: {', '.join(missing)}",
            missing_criteria=missing,
            diagnostic_logs=logs
        )
        logger.info(f"🔎 [COGNITIVE-VERIFIER]: Passed={passed}, Classification={fail_cls.value}, Summary='{ver_res.summary}'")
        return ver_res

    @classmethod
    def create_experience(
        cls,
        request: CognitiveRequest,
        verifier_result: VerificationResult,
        strategy: str,
        tools_used: List[str],
        failure_reason: Optional[str] = None
    ) -> ExperienceRecord:
        """
        Tạo Hồ sơ Trải nghiệm (ExperienceRecord) bao gồm Bài học tiêu cực qua từng Attempt.
        """
        negative_lessons: List[str] = []
        if not verifier_result.passed:
            for miss in verifier_result.missing_criteria:
                if "EXCEL_CORRUPTED" in miss:
                    negative_lessons.append("Avoid raw file stream writing for Excel. Use openpyxl.Workbook.save().")
                elif "PHYSICAL_FILE_MISSING" in miss:
                    negative_lessons.append("Ensure target directory exists before running script.")

        rec = ExperienceRecord(
            identity=verifier_result.identity,
            task_signature=f"{request.intent}_{request.deliverable.format}",
            context_summary=request.goal[:100],
            strategy_used=strategy,
            tools_used=tools_used,
            model_profile_used="AMG_SELECTED_MODEL",
            outcome="SUCCESS" if verifier_result.passed else "FAILED",
            failure_classification=verifier_result.failure_classification,
            failure_cause=failure_reason or verifier_result.summary if not verifier_result.passed else None,
            recovery_action=verifier_result.recommended_recovery.value if not verifier_result.passed else None,
            negative_lessons=negative_lessons,
            confidence_rating=verifier_result.score
        )
        logger.info(f"🧠 [ENGRAM-EXPERIENCE]: Logged record attempt={rec.identity.attempt_id}, outcome={rec.outcome}")
        return rec
